import base64
import copy
import io
import json
import logging
import re
from collections import Counter
from itertools import count
from typing import Dict, List, Tuple

from odoo import Command, _, fields, models
from odoo.exceptions import UserError

from ..models.l10n_ec_chart_template import COMPANY_PLACEHOLDER

_logger = logging.getLogger(__name__)

GROUP_SHEET = "Account Groups"
ACCOUNT_SHEET = "Accounts"
REFERENCE_TEMPLATE_CODE = "ec"


class L10nEcCoaImportWizard(models.TransientModel):
    _name = "l10n.ec.coa.import.wizard"
    _description = "Import Ecuadorian COA"

    RECONCILABLE_TYPES = {
        "asset_cash",
        "asset_receivable",
        "liability_credit_card",
        "liability_payable",
    }

    state = fields.Selection(
        [
            ("upload", "Upload"),
            ("review", "Review"),
            ("configure", "Configure"),
            ("tax_mapping", "Tax Account Mapping"),
            ("done", "Done"),
        ],
        string="State",
        readonly=True,
        default="upload",
    )
    company_id = fields.Many2one(
        "res.company",
        string="Company",
        required=True,
        default=lambda self: self.env.company.id,
    )
    data_file = fields.Binary(string="COA Template", required=True)
    data_filename = fields.Char(string="Filename")
    summary_message = fields.Text(readonly=True)
    parsed_group_payload = fields.Text(readonly=True)
    parsed_account_payload = fields.Text(readonly=True)
    account_preview = fields.Text(readonly=True)
    warning_message = fields.Text(readonly=True)
    force_override = fields.Boolean(
        string="Override Existing Chart",
        help="If enabled, existing chart data (accounts, groups, template assignments) will be cleared before importing the uploaded file.",
    )
    generated_template_id = fields.Many2one(
        "l10n.ec.chart.template",
        string="Generated Template",
        readonly=True,
    )

    # Configuration fields (filled at configure step)
    default_receivable_account_id = fields.Many2one(
        "account.account",
        string="Default Receivable",
    )
    default_payable_account_id = fields.Many2one(
        "account.account",
        string="Default Payable",
    )
    default_tax_purchase_id = fields.Many2one(
        "account.tax",
        string="Default Purchase Tax",
    )
    default_tax_sale_id = fields.Many2one(
        "account.tax",
        string="Default Sales Tax",
    )
    account_default_pos_receivable_account_id = fields.Many2one(
        "account.account",
        string="POS Default Receivable",
    )
    income_currency_exchange_account_id = fields.Many2one(
        "account.account",
        string="Currency Exchange Gain",
    )
    expense_currency_exchange_account_id = fields.Many2one(
        "account.account",
        string="Currency Exchange Loss",
    )
    account_journal_early_pay_discount_loss_account_id = fields.Many2one(
        "account.account",
        string="Early Payment Discount Loss",
    )
    account_journal_early_pay_discount_gain_account_id = fields.Many2one(
        "account.account",
        string="Early Payment Discount Gain",
    )
    default_cash_difference_income_account_id = fields.Many2one(
        "account.account",
        string="Cash Difference Income",
    )
    default_cash_difference_expense_account_id = fields.Many2one(
        "account.account",
        string="Cash Difference Expense",
    )
    
    # Tax account mapping fields
    tax_account_mapping_data = fields.Text(
        string="Tax Account Mapping Data",
        help="JSON data storing the mapping between original tax account codes and new imported account IDs",
    )
    tax_account_mapping_line_ids = fields.One2many(
        "l10n.ec.tax.account.mapping.line",
        "wizard_id",
        string="Tax Account Mappings",
    )

    def action_parse_file(self):
        self.ensure_one()
        if self.state != "upload":
            return self._reopen()
        if not self.data_file:
            raise UserError(_("Please upload the XLSX template."))
        groups, accounts = self._parse_template()
        if not groups:
            raise UserError(_("No account groups found in sheet '%s'.", GROUP_SHEET))
        if not accounts:
            raise UserError(_("No accounts found in sheet '%s'.", ACCOUNT_SHEET))
        type_labels = dict(self.env["account.account"]._fields["account_type"].selection)
        type_counts = Counter(
            account.get("account_type") or "unknown" for account in accounts
        )
        preview_lines = []
        if type_counts:
            preview_lines.append(_("Account type breakdown:"))
            for account_type, count in sorted(
                type_counts.items(),
                key=lambda item: (-item[1], type_labels.get(item[0], item[0])),
            ):
                label = type_labels.get(account_type, account_type)
                preview_lines.append("- %s: %s" % (label, count))
        else:
            preview_lines.append(_("No accounts detected."))
        if len(accounts) > 0:
            preview_lines.append(
                _(
                    "Detailed account preview is hidden. Download the template if you need to review the full data."
                )
            )
        summary = _("%s groups and %s leaf accounts ready to import. Parent accounts will be converted to groups automatically.") % (
            len(groups),
            len(accounts),
        )
        warning_message = self._compute_review_warning(self.company_id) or False
        self.write(
            {
                "generated_template_id": False,
                "parsed_group_payload": json.dumps(groups, ensure_ascii=False, indent=2),
                "parsed_account_payload": json.dumps(
                    accounts, ensure_ascii=False, indent=2
                ),
                "account_preview": "\n".join(preview_lines),
                "summary_message": summary,
                "warning_message": warning_message,
                "state": "review",
            }
        )
        return self._reopen()

    def action_import(self):
        self.ensure_one()
        if self.state != "review":
            return self._reopen()
        groups, accounts = self._parsed_payloads()
        company = self.company_id
        force_override = bool(self.force_override)
        if force_override:
            self._validate_company_readiness(company, True)
            
        template = self._create_chart_template(company, groups, accounts)
        
        # Install the chart template immediately if force_override is enabled
        # This creates accounts, groups, and taxes (with placeholder tax accounts)
        if force_override:
            self._install_chart_template(company, template)
            summary = _(
                "Chart of accounts installed for %s with %s groups and %s accounts. Review defaults below.",
                company.display_name,
                len(groups),
                len(accounts),
            )
        else:
            summary = _(
                "Registered template '%s' with %s groups and %s accounts. Install it later from Accounting Settings.",
                template.name,
                len(groups),
                len(accounts),
            )
        
        defaults, warnings = self._suggest_company_defaults(company, template)
        warning_message = "\n".join(warnings) if warnings else False
        self.write(
            {
                "state": "configure",
                "summary_message": summary,
                "warning_message": warning_message,
                "generated_template_id": template.id,
                "force_override": False,
                "default_receivable_account_id": (
                    defaults.get("receivable").id if defaults.get("receivable") else False
                ),
                "default_payable_account_id": (
                    defaults.get("payable").id if defaults.get("payable") else False
                ),
                "default_tax_sale_id": (
                    defaults.get("sale_tax").id if defaults.get("sale_tax") else False
                ),
                "default_tax_purchase_id": (
                    defaults.get("purchase_tax").id if defaults.get("purchase_tax") else False
                ),
                "account_default_pos_receivable_account_id": (
                    defaults.get("pos_receivable").id if defaults.get("pos_receivable") else False
                ),
                "income_currency_exchange_account_id": (
                    defaults.get("income_exchange").id if defaults.get("income_exchange") else False
                ),
                "expense_currency_exchange_account_id": (
                    defaults.get("expense_exchange").id if defaults.get("expense_exchange") else False
                ),
                "account_journal_early_pay_discount_loss_account_id": (
                    defaults.get("early_pay_loss").id if defaults.get("early_pay_loss") else False
                ),
                "account_journal_early_pay_discount_gain_account_id": (
                    defaults.get("early_pay_gain").id if defaults.get("early_pay_gain") else False
                ),
                "default_cash_difference_income_account_id": (
                    defaults.get("cash_diff_income").id if defaults.get("cash_diff_income") else False
                ),
                "default_cash_difference_expense_account_id": (
                    defaults.get("cash_diff_expense").id if defaults.get("cash_diff_expense") else False
                ),
            }
        )
        return self._reopen()

    def _parsed_payloads(self):
        try:
            group_payload = json.loads(self.parsed_group_payload or "[]")
            account_payload = json.loads(self.parsed_account_payload or "[]")
        except json.JSONDecodeError as err:
            raise UserError(
                _("Unable to decode stored payloads. Please restart the wizard and upload the file again. (%s)")
                % err
            ) from err
        if not isinstance(group_payload, list) or not isinstance(account_payload, list):
            raise UserError(
                _("Stored payloads are corrupted. Please restart the wizard and upload a valid file.")
            )
        return group_payload, account_payload

    def action_apply_configuration(self):
        self.ensure_one()
        if self.state != "configure":
            return self._reopen()
        company = self.company_id
        self._apply_company_defaults(company)
        
        # Prepare tax account mapping
        base_data = self._get_reference_template_data(company)
        tax_accounts = self._extract_tax_accounts_from_template(base_data)
        
        if not tax_accounts:
            # No tax accounts to map, skip to done
            self.write({
                "state": "done",
                "summary_message": _(
                    "Custom chart of accounts installed for %s. No tax accounts found to map.",
                    company.display_name
                ),
            })
        else:
            # Create mapping lines with suggestions
            mapping_lines = []
            for code, name in sorted(tax_accounts.items()):
                suggested_account = self._suggest_account_for_tax_mapping(company, code, name)
                mapping_lines.append(Command.create({
                    "original_code": code,
                    "original_name": name,
                    "new_account_id": suggested_account.id if suggested_account else False,
                }))
            
            self.write({
                "state": "tax_mapping",
                "tax_account_mapping_line_ids": mapping_lines,
                "summary_message": _(
                    "Please review and adjust the %d suggested tax account mapping(s) below.",
                    len(tax_accounts)
                ),
            })
        
        return self._reopen()

    def _suggest_account_for_tax_mapping(self, company, original_code, original_name):
        """
        Suggest a matching account from the imported COA for a tax account mapping.
        
        Tries multiple strategies:
        1. Exact code match
        2. Similar code match (numeric prefix)
        3. Similar name match (fuzzy)
        
        Args:
            company: res.company record
            original_code: str - original account code
            original_name: str - original account name
            
        Returns:
            account.account record or empty recordset
        """
        Account = self.env["account.account"].sudo().with_company(company).with_context(active_test=False)
        
        # Strategy 1: Exact code match
        account = Account.search([
            ("company_ids", "in", company.ids),
            ("code", "=", original_code),
        ], limit=1)
        
        if account:
            _logger.info(
                "Tax account suggestion: exact code match for %s -> %s",
                original_code,
                account.code
            )
            return account
        
        # Strategy 2: Similar code match (try numeric prefix)
        if original_code.isdigit() and len(original_code) > 2:
            # Try progressively shorter prefixes
            for length in range(len(original_code), 1, -1):
                prefix = original_code[:length]
                account = Account.search([
                    ("company_ids", "in", company.ids),
                    ("code", "=like", f"{prefix}%"),
                ], limit=1, order="code")
                
                if account:
                    _logger.info(
                        "Tax account suggestion: prefix match for %s -> %s (prefix: %s)",
                        original_code,
                        account.code,
                        prefix
                    )
                    return account
        
        # Strategy 3: Name similarity
        if original_name:
            # Normalize name for comparison
            normalized_original = original_name.lower().strip()
            
            # Search for accounts with similar names
            all_accounts = Account.search([
                ("company_ids", "in", company.ids),
            ])
            
            best_match = None
            best_score = 0
            
            for acc in all_accounts:
                normalized_acc_name = acc.name.lower().strip()
                
                # Simple word overlap scoring
                original_words = set(normalized_original.split())
                acc_words = set(normalized_acc_name.split())
                
                if original_words and acc_words:
                    overlap = len(original_words & acc_words)
                    score = overlap / max(len(original_words), len(acc_words))
                    
                    if score > best_score and score > 0.3:  # At least 30% match
                        best_score = score
                        best_match = acc
            
            if best_match:
                _logger.info(
                    "Tax account suggestion: name match for '%s' -> %s '%s' (score: %.2f)",
                    original_name,
                    best_match.code,
                    best_match.name,
                    best_score
                )
                return best_match
        
        _logger.warning(
            "No suitable tax account suggestion found for %s (%s)",
            original_code,
            original_name
        )
        return Account  # Return empty recordset

    def action_apply_tax_mapping(self):
        """Apply tax account mappings and finalize the import."""
        self.ensure_one()
        if self.state != "tax_mapping":
            return self._reopen()
        
        # Validate mappings from One2many lines
        unmapped_lines = self.tax_account_mapping_line_ids.filtered(lambda line: not line.new_account_id)
        
        if unmapped_lines:
            unmapped_list = "\n".join([
                f"{line.original_code} ({line.original_name})"
                for line in unmapped_lines
            ])
            raise UserError(
                _("The following tax accounts must be mapped before proceeding:\n\n%s")
                % unmapped_list
            )
        
        # Build mappings dictionary from lines
        mappings = {
            line.original_code: line.new_account_id.id
            for line in self.tax_account_mapping_line_ids
            if line.new_account_id
        }
        
        # Apply mappings to the generated template AND update installed taxes
        company = self.company_id
        
        _logger.info(
            "Starting tax account mapping for company %s with %d mappings",
            company.display_name,
            len(mappings)
        )
        
        self._apply_tax_account_mappings_to_installed_taxes(company, mappings)
        
        self.write({
            "state": "done",
            "summary_message": _(
                "Import completed successfully! Tax accounts updated with %d mappings.",
                len(mappings)
            ),
        })
        
        return self._reopen()
    
    def _apply_tax_account_mappings_to_template(self, mappings):
        """Update the generated template's tax records with new account mappings."""
        if not self.generated_template_id or not mappings:
            return
        
        template = self.generated_template_id
        payload = copy.deepcopy(template.payload or {})
        
        # Get reference data to identify original account xmlids
        company = self.company_id
        base_data = self._get_reference_template_data(company)
        reference_account_codes = self._extract_reference_account_codes(base_data)
        
        # Create reverse mapping: xmlid -> new account code
        code_to_xmlid = {code: xmlid for xmlid, code in reference_account_codes.items()}
        
        # Get account map from payload
        account_data = payload.get("account.account", {})
        account_code_to_xmlid = {}
        for xmlid, values in account_data.items():
            if "code" in values:
                account_code_to_xmlid[values["code"]] = xmlid
        
        # Build xmlid mapping
        xmlid_mappings = {}
        for orig_code, new_account_id in mappings.items():
            orig_xmlid = code_to_xmlid.get(orig_code)
            if orig_xmlid and new_account_id:
                new_account = self.env["account.account"].browse(int(new_account_id))
                new_code = new_account.code
                new_xmlid = account_code_to_xmlid.get(new_code)
                if new_xmlid:
                    xmlid_mappings[orig_xmlid] = new_xmlid
        
        # Apply mappings to tax records
        taxes = payload.get("account.tax", {})
        for tax_xmlid, tax_values in (taxes or {}).items():
            # Map cash basis accounts
            for field in ("cash_basis_account_id", "cash_basis_base_account_id", "cash_basis_transition_account_id"):
                if field in tax_values and tax_values[field]:
                    orig_xmlid = tax_values[field]
                    if orig_xmlid in xmlid_mappings:
                        tax_values[field] = xmlid_mappings[orig_xmlid]
            
            # Map repartition line accounts
            for rep_field in ("repartition_line_ids", "invoice_repartition_line_ids", "refund_repartition_line_ids"):
                if rep_field in tax_values:
                    updated_commands = []
                    for command in tax_values[rep_field] or []:
                        if isinstance(command, tuple) and len(command) >= 3 and isinstance(command[2], dict):
                            line_values = dict(command[2])
                            if "account_id" in line_values and line_values["account_id"]:
                                orig_xmlid = line_values["account_id"]
                                if orig_xmlid in xmlid_mappings:
                                    line_values["account_id"] = xmlid_mappings[orig_xmlid]
                            updated_commands.append((command[0], command[1], line_values))
                        else:
                            updated_commands.append(command)
                    tax_values[rep_field] = updated_commands
        
        # Apply mappings to tax group records
        tax_groups = payload.get("account.tax.group", {})
        for group_xmlid, group_values in (tax_groups or {}).items():
            for field in ("tax_payable_account_id", "tax_receivable_account_id"):
                if field in group_values and group_values[field]:
                    orig_xmlid = group_values[field]
                    if orig_xmlid in xmlid_mappings:
                        group_values[field] = xmlid_mappings[orig_xmlid]
        
        # Save updated payload
        template.write({"payload": payload})
        
        _logger.info(
            "Applied %d tax account mappings to template %s",
            len(xmlid_mappings),
            template.name
        )

    def _apply_tax_account_mappings_to_installed_taxes(self, company, mappings):
        """
        Update already-installed tax records with new account mappings.
        This works by matching the template definition to find which taxes/lines 
        should use which accounts, then applying the user's account selections.
        
        Args:
            company: res.company record
            mappings: dict {original_template_code: new_account_id}
        """
        if not mappings:
            _logger.info("No tax account mappings to apply")
            return
        
        _logger.info(
            "Tax account mapping for company %s: processing %d account mappings",
            company.display_name,
            len(mappings)
        )
        
        # Get the reference template to know which taxes should use which accounts
        base_data = self._get_reference_template_data(company)
        reference_account_codes = self._extract_reference_account_codes(base_data)
        base_taxes = base_data.get("account.tax", {})
        base_tax_groups = base_data.get("account.tax.group", {})
        
        # Build xmlid -> new_account_id mapping
        xmlid_to_new_account_id = {}
        for xmlid, code in reference_account_codes.items():
            if code in mappings:
                xmlid_to_new_account_id[xmlid] = mappings[code]
        
        _logger.info("Mapped %d xmlids to new accounts", len(xmlid_to_new_account_id))
        
        # Log sample of what we're looking for
        _logger.info("Processing %d taxes from template", len(base_taxes))
        sample_names = []
        sample_xmlids = []
        for i, (tax_xmlid, tax_data) in enumerate(base_taxes.items()):
            if i < 5:
                tax_name = tax_data.get("name")
                if isinstance(tax_name, dict):
                    tax_name = tax_name.get("en_US", list(tax_name.values())[0] if tax_name else "")
                sample_names.append(tax_name)
                sample_xmlids.append(tax_xmlid)
        _logger.info("Sample tax names from template: %s", sample_names)
        _logger.info("Sample tax xmlids from template: %s", sample_xmlids)
        
        # Check what taxes actually exist
        all_taxes = self.env["account.tax"].search([("company_id", "=", company.id)])
        _logger.info("Found %d installed taxes in company %s (id=%d)", len(all_taxes), company.name, company.id)
        if all_taxes:
            _logger.info("Sample installed tax names: %s", [t.name for t in all_taxes[:5]])
        
        updated_taxes = 0
        updated_lines = 0
        updated_tax_groups = 0
        
        taxes_found = 0
        taxes_not_found = 0
        
        # Update taxes: match by xmlid from template
        for tax_xmlid, tax_data in base_taxes.items():
            # The template xmlids are like "tax_vat_15_412"
            # But installed xmlids are like "account.1_tax_vat_15_412" (module: account, name: {company_id}_{xmlid})
            
            # Extract the bare xmlid (without module prefix if present)
            bare_xmlid = tax_xmlid.split('.', 1)[1] if '.' in tax_xmlid else tax_xmlid
            
            # Try to find the tax by the installed format: account.{company_id}_{bare_xmlid}
            installed_xmlid = f"account.{company.id}_{bare_xmlid}"
            tax = self.env.ref(installed_xmlid, raise_if_not_found=False)
            
            if not tax:
                # Try the original xmlid as-is
                tax = self.env.ref(tax_xmlid, raise_if_not_found=False)
            
            if not tax:
                # Try with l10n_ec_coa_import prefix as last resort
                alternative_xmlid = f"l10n_ec_coa_import.{self.generated_template_id.code}_{bare_xmlid}" if self.generated_template_id else None
                if alternative_xmlid:
                    tax = self.env.ref(alternative_xmlid, raise_if_not_found=False)
            
            if not tax or tax.company_id != company:
                taxes_not_found += 1
                if taxes_not_found <= 3:
                    _logger.info("Tax not found for xmlid: %s (tried: %s)", tax_xmlid, installed_xmlid)
                continue
            
            taxes_found += 1
            
            # Debug: log first tax details
            if taxes_found == 1:
                _logger.info("First matched tax %s, data keys: %s", tax.name, list(tax_data.keys()))
                for field in ("cash_basis_account_id", "cash_basis_base_account_id", "cash_basis_transition_account_id"):
                    if field in tax_data:
                        _logger.info("  %s: %s (in xmlid map: %s)", field, tax_data[field], tax_data[field] in xmlid_to_new_account_id if tax_data[field] else False)
                
                # Check repartition_line_ids structure
                if "repartition_line_ids" in tax_data:
                    rep_lines = tax_data["repartition_line_ids"]
                    _logger.info("  repartition_line_ids: %d lines", len(rep_lines) if rep_lines else 0)
                    if rep_lines and len(rep_lines) > 0:
                        first_line = rep_lines[0]
                        _logger.info("    First line type: %s, content: %s", type(first_line), first_line if isinstance(first_line, dict) else first_line[:3] if isinstance(first_line, (tuple, list)) else str(first_line)[:100])
            
            # Update cash basis accounts if defined in template
            tax_values = {}
            for field in ("cash_basis_account_id", "cash_basis_base_account_id", "cash_basis_transition_account_id"):
                if field in tax_data and tax_data[field]:
                    account_xmlid = tax_data[field]
                    if account_xmlid in xmlid_to_new_account_id:
                        tax_values[field] = xmlid_to_new_account_id[account_xmlid]
            
            if tax_values:
                tax.write(tax_values)
                updated_taxes += 1
                _logger.info("Updated tax %s with accounts: %s", tax.name, list(tax_values.keys()))
            
            # Update repartition lines - handle both old and new formats
            rep_fields_to_check = []
            
            # New format: separate invoice and refund fields
            if "invoice_repartition_line_ids" in tax_data or "refund_repartition_line_ids" in tax_data:
                rep_fields_to_check = [
                    ("invoice_repartition_line_ids", tax.invoice_repartition_line_ids),
                    ("refund_repartition_line_ids", tax.refund_repartition_line_ids)
                ]
            # Old format: combined repartition_line_ids
            elif "repartition_line_ids" in tax_data:
                # In the old format, all lines are in one field
                # We need to split them by document_type or use_in_tax_closing
                all_lines = tax.invoice_repartition_line_ids | tax.refund_repartition_line_ids
                rep_fields_to_check = [("repartition_line_ids", all_lines)]
            
            for rep_field, actual_lines in rep_fields_to_check:
                if rep_field not in tax_data:
                    continue
                
                template_lines = tax_data[rep_field]
                if not template_lines:
                    continue
                
                # Debug: log first tax repartition details
                if taxes_found == 1 and rep_field in ("invoice_repartition_line_ids", "repartition_line_ids"):
                    _logger.info("  %s has %d template lines", rep_field, len(template_lines))
                    for i, line_cmd in enumerate(template_lines[:2]):
                        if isinstance(line_cmd, (tuple, list)) and len(line_cmd) >= 3:
                            line_data = line_cmd[2]
                            if isinstance(line_data, dict):
                                _logger.info("    Line %d: account_id=%s, factor=%s, type=%s", 
                                           i, line_data.get("account_id"), 
                                           line_data.get("factor_percent"), 
                                           line_data.get("repartition_type"))
                
                # Get the actual installed lines - they're already provided in rep_fields_to_check
                
                # Match template lines to actual lines by position and type
                for i, line_cmd in enumerate(template_lines):
                    if not isinstance(line_cmd, (tuple, list)) or len(line_cmd) < 3:
                        continue
                    
                    line_data = line_cmd[2]
                    if not isinstance(line_data, dict):
                        continue
                    
                    # Find matching actual line by factor and type
                    factor = line_data.get("factor_percent", 100.0)
                    rep_type = line_data.get("repartition_type", "tax")
                    doc_type = line_data.get("document_type")  # 'invoice' or 'refund' in the old format
                    
                    # Filter actual_lines by document_type if present
                    lines_to_search = actual_lines
                    if doc_type and rep_field == "repartition_line_ids":
                        # In combined format, filter by document_type
                        if doc_type == "invoice":
                            lines_to_search = tax.invoice_repartition_line_ids
                        elif doc_type == "refund":
                            lines_to_search = tax.refund_repartition_line_ids
                    
                    # Try to find exact match
                    matching = lines_to_search.filtered(
                        lambda l: abs(l.factor_percent - factor) < 0.01 
                        and l.repartition_type == rep_type
                    )
                    
                    if not matching:
                        _logger.warning(
                            "No matching line found for tax %s (factor=%s, type=%s, doc=%s) - %d lines in search set",
                            tax.name, factor, rep_type, doc_type or 'N/A', len(lines_to_search)
                        )
                        continue
                    
                    # Take first match (or match by position if multiple)
                    actual_line = matching[0] if len(matching) == 1 else matching[min(i, len(matching) - 1)]
                    
                    # Check if this line should have an account
                    if "account_id" in line_data and line_data["account_id"]:
                        account_xmlid = line_data["account_id"]
                        if account_xmlid in xmlid_to_new_account_id:
                            new_account_id = xmlid_to_new_account_id[account_xmlid]
                            if actual_line.account_id.id != new_account_id:
                                actual_line.write({"account_id": new_account_id})
                                updated_lines += 1
                                _logger.info(
                                    "Updated repartition line for tax %s (factor=%s, type=%s, doc=%s)",
                                    tax.name, factor, rep_type, doc_type or 'N/A'
                                )
                        else:
                            _logger.warning(
                                "Account xmlid '%s' not found in mapping for tax %s (factor=%s, type=%s, doc=%s)",
                                account_xmlid, tax.name, factor, rep_type, doc_type or 'N/A'
                            )
                    else:
                        # Template doesn't define an account for this line (intentional for some taxes)
                        if rep_type == "tax":  # Only log for tax lines, base lines normally don't have accounts
                            _logger.debug(
                                "Tax %s: template doesn't define account for %s line (type=%s, factor=%s, doc=%s). "
                                "This is normal for certain tax types (e.g., 512 taxes).",
                                tax.name, rep_field, rep_type, factor, doc_type or 'N/A'
                            )
        
        # Update tax groups: match by xmlid
        for group_xmlid, group_data in (base_tax_groups or {}).items():
            # Extract bare xmlid and try the account.{company_id}_{bare_xmlid} format
            bare_xmlid = group_xmlid.split('.', 1)[1] if '.' in group_xmlid else group_xmlid
            installed_xmlid = f"account.{company.id}_{bare_xmlid}"
            
            tax_group = self.env.ref(installed_xmlid, raise_if_not_found=False)
            
            if not tax_group:
                # Try original xmlid
                tax_group = self.env.ref(group_xmlid, raise_if_not_found=False)
            
            if not tax_group:
                # Try with l10n_ec_coa_import prefix
                alternative_xmlid = f"l10n_ec_coa_import.{self.generated_template_id.code}_{bare_xmlid}" if self.generated_template_id else None
                if alternative_xmlid:
                    tax_group = self.env.ref(alternative_xmlid, raise_if_not_found=False)
            
            if not tax_group:
                continue
            
            group_values = {}
            for field in ("tax_payable_account_id", "tax_receivable_account_id"):
                if field in group_data and group_data[field]:
                    account_xmlid = group_data[field]
                    if account_xmlid in xmlid_to_new_account_id:
                        group_values[field] = xmlid_to_new_account_id[account_xmlid]
            
            if group_values:
                tax_group.write(group_values)
                updated_tax_groups += 1
                _logger.info("Updated tax group %s with accounts", tax_group.name)
        
        _logger.info("Matched %d taxes, %d not found", taxes_found, taxes_not_found)
        _logger.info(
            "Tax account mapping completed: %d taxes, %d repartition lines, %d tax groups updated",
            updated_taxes,
            updated_lines,
            updated_tax_groups
        )

    def action_download_template(self):
        return {
            "type": "ir.actions.act_url",
            "url": "/l10n_ec_coa_import/template/download",
            "target": "self",
        }

    def _reopen(self):
        """Return the wizard window action to reload current record."""
        return {
            "type": "ir.actions.act_window",
            "res_model": self._name,
            "res_id": self.id,
            "view_mode": "form",
            "view_id": self.env.ref(
                "l10n_ec_coa_import.view_l10n_ec_coa_import_wizard"
            ).id,
            "target": "new",
        }

    # ------------------------------------------------------------------
    # Parsing helpers
    # ------------------------------------------------------------------
    def _parse_template(self) -> Tuple[List[Dict], List[Dict]]:
        """
        Parse the uploaded XLSX COA template and extract account groups and accounts.

        Returns:
            Tuple[List[Dict], List[Dict]]: A tuple containing a list of group dictionaries and a list of account dictionaries.
        """
        try:
            from openpyxl import load_workbook
        except ImportError as err:  # pragma: no cover
            raise UserError(
                _(
                    "openpyxl is required to parse XLSX files. Please install it or contact your administrator."
                )
            ) from err

        decoding = base64.b64decode(self.data_file)
        try:
            workbook = load_workbook(io.BytesIO(decoding), read_only=False, data_only=True)
        except Exception as err:
            raise UserError(_("Unable to read the XLSX file: %s") % err) from err

        if ACCOUNT_SHEET not in workbook.sheetnames:
            raise UserError(
                _("Sheet '%s' is missing from the template.") % ACCOUNT_SHEET
            )

        groups = []
        if GROUP_SHEET in workbook.sheetnames:
            groups = self._extract_group_rows(workbook[GROUP_SHEET])
        accounts = self._extract_account_rows(workbook[ACCOUNT_SHEET])
        groups, accounts = self._post_process_rows(groups, accounts)
        if not groups:
            raise UserError(
                _(
                    "No account groups could be inferred. Please review the uploaded file."
                )
            )
        if not accounts:
            raise UserError(
                _("No accounts found in sheet '%s'.") % ACCOUNT_SHEET
            )
        return groups, accounts

    def _read_headers(self, sheet) -> List[str]:
        header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), [])
        headers = []
        seen = set()
        for cell in header_row:
            label = str(cell or "").strip().lower().replace(" ", "_")
            label = re.sub(r"[^0-9a-z_]+", "", label)
            if not label:
                continue
            if label in seen:
                raise UserError(
                    _("Duplicate column '%s' detected in sheet '%s'.", label, sheet.title)
                )
            headers.append(label)
            seen.add(label)
        if not headers:
            raise UserError(_("No headers found in sheet '%s'.", sheet.title))
        return headers

    def _iterate_sheet(self, sheet, headers):
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if not any(value not in (None, "", False) for value in row):
                continue
            values = {}
            for index, header in enumerate(headers):
                current = row[index] if index < len(row) else None
                if isinstance(current, str):
                    current = current.strip()
                values[header] = current
            yield values

    def _extract_group_rows(self, sheet) -> List[Dict]:
        headers = self._read_headers(sheet)
        required = {"code", "name"}
        missing = required - set(headers)
        if missing:
            raise UserError(
                _(
                    "Columns %s are required in sheet '%s'.",
                    ", ".join(sorted(missing)),
                    GROUP_SHEET,
                )
            )
        rows = []
        for row_values in self._iterate_sheet(sheet, headers):
            if not row_values.get("code"):
                continue
            prefix_start = self._normalize_group_prefix_value(
                row_values.get("code_prefix_start")
            )
            prefix_end = self._normalize_group_prefix_value(
                row_values.get("code_prefix_end")
            )
            if prefix_start and not prefix_end:
                prefix_end = prefix_start
            elif prefix_end and not prefix_start:
                prefix_start = prefix_end
            rows.append(
                {
                    "code": str(row_values.get("code")).strip(),
                    "name": str(row_values.get("name")).strip(),
                    "parent_code": (row_values.get("parent_code") or "").strip(),
                    "code_prefix_start": prefix_start,
                    "code_prefix_end": prefix_end,
                    "sequence": int(row_values.get("sequence") or 10),
                }
            )
        return rows

    def _extract_account_rows(self, sheet) -> List[Dict]:
        headers = self._read_headers(sheet)
        headers_set = set(headers)
        required = {"code", "name", "account_type"}
        missing = required - set(headers)
        if missing:
            raise UserError(
                _(
                    "Columns %s are required in sheet '%s'.",
                    ", ".join(sorted(missing)),
                    ACCOUNT_SHEET,
                )
            )
        rows = []
        for row_values in self._iterate_sheet(sheet, headers):
            code = str(row_values.get("code") or "").strip()
            if not code:
                continue
            reconcile = None
            if "reconcile" in headers_set:
                reconcile = bool(row_values.get("reconcile"))
            rows.append(
                {
                    "code": code,
                    "name": str(row_values.get("name") or "").strip(),
                    "account_type": str(row_values.get("account_type") or "").strip(),
                    "group_code": str(row_values.get("group_code") or "").strip(),
                    "reconcile": reconcile,
                    "deprecated": bool(row_values.get("deprecated")) if "deprecated" in headers_set else False,
                }
            )
        return rows

    def _post_process_rows(
        self, groups: List[Dict], accounts: List[Dict]
    ) -> Tuple[List[Dict], List[Dict]]:
        if not accounts:
            return groups, accounts

        group_index = {group["code"]: group for group in groups}
        if not group_index:
            group_index = self._infer_groups_from_accounts(accounts)
            groups = sorted(group_index.values(), key=lambda item: item["sequence"])
        else:
            for account in accounts:
                if not account.get("group_code"):
                    account["group_code"] = self._match_group_code(
                        account["code"], group_index
                    )
                    if not account["group_code"]:
                        raise UserError(
                            _("Unable to infer a group for account %s.")
                            % account["code"]
                        )
                if account["group_code"] not in group_index:
                    raise UserError(
                        _(
                            "Account %s references non-existing group %s.",
                            account["code"],
                            account["group_code"],
                        )
                    )
                # Note: reconcile and deprecated will be set after transformation
                pass

        # Transform accounts with children into groups
        groups, accounts = self._transform_parent_accounts_to_groups(groups, accounts)
        
        # Now process account properties for the remaining leaf accounts
        for account in accounts:
            account["reconcile"] = self._normalize_reconcile(account)
            account["deprecated"] = bool(account.get("deprecated"))
        
        self._validate_duplicates(groups, accounts)
        return groups, accounts

    def _transform_parent_accounts_to_groups(self, groups: List[Dict], accounts: List[Dict]) -> Tuple[List[Dict], List[Dict]]:
        """
        Transform accounts that have children into account groups, keeping only leaf accounts as actual accounts.
        
        Returns:
            Tuple[List[Dict], List[Dict]]: Updated groups and accounts lists
        """
        if not accounts:
            return groups, accounts
            
        # Build account hierarchy
        account_index = {acc["code"]: acc for acc in accounts}
        children_map = {}
        parent_map = {}
        
        # Find parent-child relationships using a more sophisticated approach
        sorted_codes = sorted(account_index.keys(), key=lambda x: (len(x), x))
        
        for account in accounts:
            code = account["code"]
            children_map[code] = []
            parent_code = None
            
            # Try different strategies to find parent
            # Strategy 1: Direct prefix match (most common)
            for potential_parent_code in sorted_codes:
                if (potential_parent_code != code 
                    and code.startswith(potential_parent_code)
                    and len(potential_parent_code) < len(code)):
                    # Ensure it's a logical parent (not just any prefix)
                    remaining = code[len(potential_parent_code):]
                    if remaining.isdigit() or len(remaining) <= 3:  # Reasonable child suffix
                        if not parent_code or len(potential_parent_code) > len(parent_code):
                            parent_code = potential_parent_code
            
            # Strategy 2: For numeric codes, try progressive shortening
            if not parent_code and code.isdigit():
                for length in range(len(code) - 1, 0, -1):
                    potential_parent = code[:length]
                    if potential_parent in account_index:
                        parent_code = potential_parent
                        break
            
            if parent_code:
                parent_map[code] = parent_code
                children_map.setdefault(parent_code, []).append(code)
        
        # Identify accounts that have children (these become groups)
        accounts_with_children = {code for code, children in children_map.items() if children}
        
        # Convert existing groups to a dict for easier manipulation
        group_index = {group["code"]: group for group in groups}
        
        # Transform parent accounts into groups
        accounts_to_remove = []
        for account in accounts:
            if account["code"] in accounts_with_children:
                # This account has children, convert it to a group
                parent_code = parent_map.get(account["code"])
                group_parent = None
                
                # Find group parent (could be an existing group or another parent account)
                if parent_code:
                    if parent_code in group_index:
                        group_parent = parent_code
                    elif parent_code in accounts_with_children:
                        group_parent = parent_code  # Will be converted to group too
                
                # Create the group
                group_index[account["code"]] = {
                    "code": account["code"],
                    "name": account["name"],
                    "parent_code": group_parent or False,
                    "code_prefix_start": account["code"],
                    "code_prefix_end": account["code"],
                    "sequence": 10,
                }
                
                # Mark account for removal
                accounts_to_remove.append(account)
        
        # Remove converted accounts
        remaining_accounts = [acc for acc in accounts if acc not in accounts_to_remove]
        
        # Update group references for remaining accounts
        for account in remaining_accounts:
            # Find the immediate parent group
            parent_code = parent_map.get(account["code"])
            while parent_code and parent_code not in group_index:
                parent_code = parent_map.get(parent_code)
            
            if parent_code:
                account["group_code"] = parent_code
            elif account.get("group_code") and account["group_code"] not in group_index:
                # Try to find a suitable group
                account["group_code"] = self._match_group_code(account["code"], group_index) or False
        
        # Convert group_index back to list and assign sequences
        updated_groups = list(group_index.values())
        self._assign_group_sequences({g["code"]: g for g in updated_groups})
        
        # Validation: ensure we don't lose any codes
        original_codes = {acc["code"] for acc in accounts}
        remaining_codes = {acc["code"] for acc in remaining_accounts}
        group_codes = {grp["code"] for grp in updated_groups}
        transformed_codes = remaining_codes | group_codes
        
        if not original_codes.issubset(transformed_codes):
            missing_codes = original_codes - transformed_codes
            raise UserError(
                _("Lost account codes during transformation: %s") % ", ".join(sorted(missing_codes))
            )
        
        _logger.info(
            "Account transformation completed: %d parent accounts → groups, %d leaf accounts → accounts, %d total groups",
            len(accounts_to_remove),
            len(remaining_accounts),
            len(updated_groups)
        )
        
        return updated_groups, remaining_accounts

    def _normalize_group_prefix_value(self, value):
        if value in (None, False):
            return False
        text = str(value).strip()
        if not text:
            return False
        text = text.replace(" ", "")
        return text

    # ------------------------------------------------------------------
    # Group inference helpers
    # ------------------------------------------------------------------

    def _infer_groups_from_accounts(self, accounts: List[Dict]):
        if not accounts:
            return {}
        separator = self._detect_code_separator(accounts)
        if separator:
            return self._infer_groups_with_separator(accounts, separator)
        return self._infer_groups_numeric(accounts)

    def _detect_code_separator(self, accounts: List[Dict]):
        separator_counts = Counter()
        for account in accounts:
            code = account.get("code") or ""
            for candidate in (".", "-"):
                if candidate in code:
                    separator_counts[candidate] += 1
        if not separator_counts:
            return False
        return separator_counts.most_common(1)[0][0]

    def _infer_groups_with_separator(self, accounts: List[Dict], separator: str):
        groups: Dict[str, Dict] = {}
        for account in sorted(accounts, key=lambda item: item["code"]):
            code = account["code"]
            segments = [segment for segment in code.split(separator) if segment]
            if len(segments) >= 2:
                for depth in range(1, len(segments)):
                    current = separator.join(segments[:depth])
                    parent = separator.join(segments[: depth - 1]) if depth > 1 else False
                    self._ensure_group_entry(groups, current, parent)
                account["group_code"] = separator.join(segments[:-1])
            else:
                account["group_code"] = False
            account["reconcile"] = self._normalize_reconcile(account)
            account["deprecated"] = bool(account.get("deprecated"))
        if groups:
            self._assign_group_sequences(groups)
        return groups

    def _infer_groups_numeric(self, accounts: List[Dict]):
        groups: Dict[str, Dict] = {}
        codes = sorted({account["code"] for account in accounts if account.get("code")}, key=lambda code: (len(code), code))
        if not codes:
            return groups
        children_map = {code: set() for code in codes}
        parent_map = {}
        for code in codes:
            parent = False
            for length in range(len(code) - 1, 0, -1):
                prefix = code[:length]
                if prefix in children_map:
                    parent = prefix
                    break
            parent_map[code] = parent
            if parent:
                children_map[parent].add(code)

        group_codes = {code for code, children in children_map.items() if children}

        ensured = set()

        def ensure_group_entry(code):
            if code in ensured:
                return
            parent_code = parent_map.get(code)
            parent_ref = parent_code if parent_code in group_codes else False
            if parent_ref:
                ensure_group_entry(parent_ref)
            self._ensure_group_entry(groups, code, parent_ref)
            ensured.add(code)

        for group_code in sorted(group_codes, key=lambda value: (len(value), value)):
            ensure_group_entry(group_code)

        for account in accounts:
            code = account["code"]
            parent_code = parent_map.get(code)
            account["group_code"] = parent_code if parent_code in group_codes else False
            account["reconcile"] = self._normalize_reconcile(account)
            account["deprecated"] = bool(account.get("deprecated"))

        if groups:
            self._assign_group_sequences(groups)
        return groups

    def _match_group_code(self, account_code, group_index):
        if not account_code:
            return False
        candidates = []
        for code, group in group_index.items():
            if not code:
                continue
            if account_code.startswith(code):
                candidates.append((len(code), code))
        if not candidates:
            return False
        candidates.sort(reverse=True)
        return candidates[0][1]

    def _normalize_reconcile(self, account: Dict) -> bool:
        account_type = account.get("account_type")
        if not account_type:
            raise UserError(
                _("Account %s is missing an account type.", account.get("code"))
            )
        explicit = account.get("reconcile")
        if explicit is None:
            return self._should_reconcile(account_type)
        return bool(explicit)

    def _should_reconcile(self, account_type: str) -> bool:
        return account_type in self.RECONCILABLE_TYPES

    def _ensure_group_entry(self, groups: Dict[str, Dict], code: str, parent_code):
        if code in groups:
            if not groups[code].get("parent_code") and parent_code:
                groups[code]["parent_code"] = parent_code
            if not groups[code].get("code_prefix_start"):
                default_prefix = self._default_group_prefix(code)
                groups[code]["code_prefix_start"] = default_prefix
                groups[code]["code_prefix_end"] = default_prefix
            return
        default_prefix = self._default_group_prefix(code)
        groups[code] = {
            "code": code,
            "name": self._default_group_name(code),
            "parent_code": parent_code or False,
            "code_prefix_start": default_prefix,
            "code_prefix_end": default_prefix,
            "sequence": 10,
        }
        return


    def _assign_group_sequences(self, groups: Dict[str, Dict]):
        for index, code in enumerate(
            sorted(groups.keys(), key=self._group_sort_key), start=1
        ):
            groups[code]["sequence"] = index * 10

    def _default_group_prefix(self, code):
        return str(code)

    def _default_group_name(self, code):
        return _("Group %s") % code

    def _group_sort_key(self, code):
        numeric = "".join(ch for ch in str(code) if ch.isdigit())
        return (len(code), int(numeric) if numeric.isdigit() else code)

    def _validate_duplicates(self, groups, accounts):
        group_codes = [group["code"] for group in groups]
        account_codes = [account["code"] for account in accounts]
        dup_groups = [code for code, count in Counter(group_codes).items() if count > 1]
        dup_accounts = [code for code, count in Counter(account_codes).items() if count > 1]
        if dup_groups:
            raise UserError(
                _("Duplicate group codes detected: %s", ", ".join(sorted(dup_groups)))
            )
        if dup_accounts:
            raise UserError(
                _("Duplicate account codes detected: %s", ", ".join(sorted(dup_accounts)))
            )

    def _compute_review_warning(self, company):
        warnings = []
        account_model = (
            self.env["account.account"].sudo().with_company(company).with_context(active_test=False)
        )
        account_count = account_model.search_count(
            [
                ("company_ids", "in", company.ids),
                ("deprecated", "=", False),
            ]
        )
        if account_count:
            warnings.append(
                _(
                    "Company %s already has %s active account(s). Enable 'Override Existing Chart' to replace them.",
                    company.display_name,
                    account_count,
                )
            )
        chart_template_name = False
        if "chart_template_id" in company._fields and company.chart_template_id:
            chart_template_name = company.chart_template_id.display_name
        elif "chart_template" in company._fields and company.chart_template:
            chart_template_name = self._get_chart_template_label(company)
        if chart_template_name:
            warnings.append(
                _(
                    "Chart template '%s' is currently assigned and will be cleared during import.",
                    chart_template_name,
                )
            )
        return "\n".join(warnings)

    def _get_chart_template_label(self, company) -> str:
        field = company._fields.get("chart_template")
        if not field:
            return False
        selection = field.selection
        if callable(selection):
            selection = selection(self.env)
        selection_dict = {}
        for entry in selection or []:
            if isinstance(entry, (tuple, list)):
                if len(entry) >= 2:
                    selection_dict[entry[0]] = entry[1]
                elif len(entry) == 1:
                    selection_dict[entry[0]] = entry[0]
            else:
                selection_dict[entry] = entry
        return selection_dict.get(company.chart_template, company.chart_template)

    # ------------------------------------------------------------------
    # Creation helpers
    # ------------------------------------------------------------------
    def _validate_company_readiness(self, company, force_override):
        account_model = (
            self.env["account.account"]
            .sudo()
            .with_company(company)
            .with_context(active_test=False)
        )
        existing_accounts = account_model.search(
            [
                ("company_ids", "in", company.ids),
                ("deprecated", "=", False),
            ],
            limit=1,
        )
        if existing_accounts and not force_override:
            raise UserError(
                _(
                    "Company %s already has active accounts. Enable 'Override Existing Chart' or clean up the accounts manually before importing."
                )
                % company.display_name
            )

        move_line_model = (
            self.env["account.move.line"]
            .sudo()
            .with_company(company)
            .with_context(active_test=False)
        )
        move_line_exists = move_line_model.search(
            [
                ("company_id", "child_of", company.id),
            ],
            limit=1,
        )
        if move_line_exists:
            raise UserError(
                _(
                    "Cannot override the chart of accounts for %s because accounting entries already exist."
                )
                % company.display_name
            )

        chart_template_fields = {}
        if "chart_template_id" in company._fields and company.chart_template_id:
            chart_template_fields["chart_template_id"] = False
        if "chart_template" in company._fields and company.chart_template:
            chart_template_fields["chart_template"] = False
        if chart_template_fields:
            company.sudo().write(chart_template_fields)
            _logger.info(
                "Cleared chart template %s for company %s before custom COA import.",
                chart_template_fields,
                company.display_name,
            )

    def _create_chart_template(self, company, groups, accounts):
        template_model = self.env["l10n.ec.chart.template"].sudo()
        name = self._generate_template_name(company)
        code = template_model.generate_code(company.display_name)
        payload = self._build_template_payload(code, groups, accounts, company)
        payload.setdefault("template_data", {}).setdefault(code, {})["name"] = name
        template_vals = {
            "name": name,
            "code": code,
            "payload": payload,
            "source_filename": self.data_filename,
            "company_ids": [Command.link(company.id)],
        }
        return template_model.create(template_vals)

    def _install_chart_template(self, company, template):
        chart_template = (
            self.env["account.chart.template"].sudo().with_context(
                default_company_id=company.id,
                allowed_company_ids=[company.id],
                chart_template_load=True,
            )
        )
        chart_template.try_loading(template.code, company, install_demo=False, force_create=True)
        self._restore_account_codes(company, template.payload.get("account.account"))

    def _suggest_company_defaults(self, company, _template):
        receivable = self._suggest_default_account(company, "asset_receivable")
        payable = self._suggest_default_account(company, "liability_payable")
        sale_tax = self._suggest_default_tax(company, "sale")
        purchase_tax = self._suggest_default_tax(company, "purchase")
        reference_data = self._get_reference_template_data(company)
        reference_company_values = (reference_data or {}).get("res.company", {}).get(company.id, {}) or {}
        reference_account_codes = self._extract_reference_account_codes(reference_data)
        account_specs = [
            (
                "account_default_pos_receivable_account_id",
                "pos_receivable",
                _("POS default receivable account"),
                ["asset_receivable"],
            ),
            (
                "income_currency_exchange_account_id",
                "income_exchange",
                _("Currency exchange gain account"),
                ["income_other"],
            ),
            (
                "expense_currency_exchange_account_id",
                "expense_exchange",
                _("Currency exchange loss account"),
                ["expense_other"],
            ),
            (
                "account_journal_early_pay_discount_loss_account_id",
                "early_pay_loss",
                _("Early payment discount loss account"),
                ["expense_other"],
            ),
            (
                "account_journal_early_pay_discount_gain_account_id",
                "early_pay_gain",
                _("Early payment discount gain account"),
                ["income_other"],
            ),
            (
                "default_cash_difference_income_account_id",
                "cash_diff_income",
                _("Cash difference income account"),
                ["income_other"],
            ),
            (
                "default_cash_difference_expense_account_id",
                "cash_diff_expense",
                _("Cash difference expense account"),
                ["expense_other"],
            ),
        ]
        defaults = {
            "receivable": receivable,
            "payable": payable,
            "sale_tax": sale_tax,
            "purchase_tax": purchase_tax,
        }
        for field_name, key, _label, fallback_types in account_specs:
            defaults[key] = self._suggest_company_account_field(
                company,
                field_name,
                reference_company_values.get(field_name),
                reference_account_codes,
                fallback_types,
            )
        warnings = self._build_configuration_warnings(defaults, account_specs)
        return defaults, warnings

    def _suggest_company_account_field(
        self,
        company,
        field_name,
        reference,
        reference_account_codes,
        fallback_types,
    ):
        current = getattr(company, field_name, False)
        if current:
            return current
        code = self._resolve_reference_account_code(reference, reference_account_codes)
        if code:
            account = self._find_account_by_code(company, code)
            if account:
                return account
        for account_type in fallback_types or []:
            candidate = self._suggest_default_account(company, account_type)
            if candidate:
                return candidate
        return self.env["account.account"]

    def _resolve_reference_account_code(self, reference, reference_account_codes):
        if isinstance(reference, str):
            return reference_account_codes.get(reference)
        if isinstance(reference, dict):
            for key in ("xmlid", "xml_id", "id"):
                value = reference.get(key)
                if isinstance(value, str):
                    code = reference_account_codes.get(value)
                    if code:
                        return code
        return False

    def _find_account_by_code(self, company, code):
        if not code:
            return self.env["account.account"]
        Account = (
            self.env["account.account"].sudo().with_company(company).with_context(active_test=False)
        )
        return Account.search([
            ("company_ids", "in", company.ids),
            ("code", "=", code),
        ], limit=1)

    def _build_configuration_warnings(self, defaults, account_specs):
        warning_labels = {
            "receivable": _("Default receivable account"),
            "payable": _("Default payable account"),
            "sale_tax": _("Default sales tax"),
            "purchase_tax": _("Default purchase tax"),
        }
        for _field, key, label, _fallback in account_specs:
            warning_labels[key] = label
        missing = []
        for key, label in warning_labels.items():
            if not defaults.get(key):
                missing.append(_("Review and set %s manually.") % label)
        return missing

    def _suggest_default_account(self, company, account_type):
        account_model = (
            self.env["account.account"]
            .sudo()
            .with_company(company)
            .with_context(active_test=False)
        )
        domain = [
            ("company_ids", "in", company.ids),
            ("account_type", "=", account_type),
            ("deprecated", "=", False),
        ]
        return account_model.search(domain, limit=1)

    def _suggest_default_tax(self, company, tax_use):
        tax_model = (
            self.env["account.tax"]
            .sudo()
            .with_company(company)
            .with_context(active_test=False)
        )
        domain = [
            ("company_id", "=", company.id),
            ("type_tax_use", "=", tax_use),
        ]
        return tax_model.search(domain, limit=1)

    def _extract_tax_accounts_from_template(self, base_data):
        """
        Extract all unique account codes used in tax definitions from the reference template.
        
        Returns:
            dict: {account_code: account_name} mapping
        """
        tax_accounts = {}
        
        # Get reference account codes from base data
        reference_account_codes = self._extract_reference_account_codes(base_data)
        base_accounts = (base_data or {}).get("account.account", {})
        
        # Extract accounts from tax records
        taxes = (base_data or {}).get("account.tax", {})
        for tax_xmlid, tax_values in (taxes or {}).items():
            # Cash basis accounts
            for field in ("cash_basis_account_id", "cash_basis_base_account_id", "cash_basis_transition_account_id"):
                if field in tax_values and tax_values[field]:
                    code = reference_account_codes.get(tax_values[field])
                    if code:
                        account_info = base_accounts.get(tax_values[field], {})
                        name = account_info.get("name", code)
                        if isinstance(name, dict):
                            name = name.get("en_US", code)
                        tax_accounts[code] = name
            
            # Repartition line accounts
            for rep_field in ("repartition_line_ids", "invoice_repartition_line_ids", "refund_repartition_line_ids"):
                if rep_field in tax_values:
                    for command in tax_values[rep_field] or []:
                        if isinstance(command, tuple) and len(command) >= 3 and isinstance(command[2], dict):
                            line_values = command[2]
                            if "account_id" in line_values and line_values["account_id"]:
                                code = reference_account_codes.get(line_values["account_id"])
                                if code:
                                    account_info = base_accounts.get(line_values["account_id"], {})
                                    name = account_info.get("name", code)
                                    if isinstance(name, dict):
                                        name = name.get("en_US", code)
                                    tax_accounts[code] = name
        
        # Extract accounts from tax group records
        tax_groups = (base_data or {}).get("account.tax.group", {})
        for group_xmlid, group_values in (tax_groups or {}).items():
            for field in ("tax_payable_account_id", "tax_receivable_account_id"):
                if field in group_values and group_values[field]:
                    code = reference_account_codes.get(group_values[field])
                    if code:
                        account_info = base_accounts.get(group_values[field], {})
                        name = account_info.get("name", code)
                        if isinstance(name, dict):
                            name = name.get("en_US", code)
                        tax_accounts[code] = name
        
        return tax_accounts

    def _apply_company_defaults(self, company):
        values = {}
        account_field_values = {
            "property_account_receivable_id": self.default_receivable_account_id.id,
            "property_account_payable_id": self.default_payable_account_id.id,
            "account_default_pos_receivable_account_id": self.account_default_pos_receivable_account_id.id,
            "income_currency_exchange_account_id": self.income_currency_exchange_account_id.id,
            "expense_currency_exchange_account_id": self.expense_currency_exchange_account_id.id,
            "account_journal_early_pay_discount_loss_account_id": self.account_journal_early_pay_discount_loss_account_id.id,
            "account_journal_early_pay_discount_gain_account_id": self.account_journal_early_pay_discount_gain_account_id.id,
            "default_cash_difference_income_account_id": self.default_cash_difference_income_account_id.id,
            "default_cash_difference_expense_account_id": self.default_cash_difference_expense_account_id.id,
        }
        for field_name, value in account_field_values.items():
            if field_name in company._fields:
                values[field_name] = value
        if values:
            company.sudo().write(values)
        tax_values = {}
        if "account_sale_tax_id" in company._fields:
            tax_values["account_sale_tax_id"] = self.default_tax_sale_id.id
        if "account_purchase_tax_id" in company._fields:
            tax_values["account_purchase_tax_id"] = self.default_tax_purchase_id.id
        if tax_values:
            company.sudo().write(tax_values)

    def _build_template_payload(self, template_code, groups, accounts, company):
        base_data = self._get_reference_template_data(company)
        base_account_codes = self._extract_reference_account_codes(base_data)
        group_map, group_data = self._prepare_group_records(groups)
        account_map, account_data = self._prepare_account_records(accounts, group_map)
        template_section = self._prepare_template_section(
            template_code,
            (base_data or {}).get("template_data", {}).get(REFERENCE_TEMPLATE_CODE, {}),
            base_account_codes,
            account_map,
        )
        company_section = self._prepare_company_section(
            (base_data or {}).get("res.company", {}).get(company.id, {}),
            base_account_codes,
            account_map,
        )
        payload = {
            "template_data": {template_code: template_section},
            "res.company": {COMPANY_PLACEHOLDER: company_section},
            "account.group": group_data,
            "account.account": account_data,
        }
        extra_models = self._prepare_reference_models(
            base_data or {},
            base_account_codes,
            account_map,
            template_code,
        )
        for model_name, model_data in extra_models.items():
            if model_data:
                payload[model_name] = model_data
        return payload

    def _get_reference_template_data(self, company):
        chart_template = (
            self.env["account.chart.template"]
            .sudo()
            .with_company(company)
            .with_context(
                allowed_company_ids=[company.id],
                default_company_id=company.id,
                chart_template_load=True,
                lang="en_US",
            )
        )
        return copy.deepcopy(
            chart_template._get_chart_template_data(REFERENCE_TEMPLATE_CODE)
        )

    def _extract_reference_account_codes(self, base_data):
        accounts = (base_data or {}).get("account.account", {})
        return {
            xmlid: values.get("code")
            for xmlid, values in accounts.items()
            if values.get("code")
        }

    def _prepare_group_records(self, groups):
        group_map = {}
        group_data = {}
        used_identifiers = set()
        for entry in groups:
            xmlid = self._make_identifier("grp", entry["code"], used_identifiers)
            group_map[entry["code"]] = xmlid
        for entry in groups:
            group_data[group_map[entry["code"]]] = {
                "name": entry["name"],
                "parent_id": group_map.get(entry.get("parent_code")) or False,
                "code_prefix_start": entry.get("code_prefix_start") or False,
                "code_prefix_end": entry.get("code_prefix_end") or False,
            }
        # Avoid overlapping prefixes that trigger account.group constraints
        used_ranges = []
        for values in group_data.values():
            start = values.get("code_prefix_start")
            end = values.get("code_prefix_end")
            if not start or not end:
                continue
            start_str = str(start)
            end_str = str(end)
            length = len(start_str)
            if len(end_str) != length:
                continue
            if start_str.isdigit() and end_str.isdigit():
                current_start = int(start_str)
                current_end = int(end_str)
                conflict = any(
                    length == used_length
                    and not (current_end < used_start or current_start > used_end)
                    for used_length, used_start, used_end in used_ranges
                )
                if conflict:
                    values["code_prefix_start"] = False
                    values["code_prefix_end"] = False
                else:
                    used_ranges.append((length, current_start, current_end))
            else:
                conflict = any(
                    length == used_length and start_str == used_start and end_str == used_end
                    for used_length, used_start, used_end in used_ranges
                )
                if conflict:
                    values["code_prefix_start"] = False
                    values["code_prefix_end"] = False
                else:
                    used_ranges.append((length, start_str, end_str))
        return group_map, group_data

    def _prepare_account_records(self, accounts, group_map):
        account_map = {}
        account_data = {}
        used_identifiers = set()
        equity_unaffected_seen = False
        for entry in accounts:
            xmlid = self._make_identifier("acc", entry["code"], used_identifiers)
            account_map[entry["code"]] = xmlid
            account_type = entry["account_type"]
            if account_type == "equity_unaffected":
                if equity_unaffected_seen:
                    account_type = "equity"
                else:
                    equity_unaffected_seen = True
            account_data[xmlid] = {
                "code": entry["code"],
                "name": entry["name"],
                "account_type": account_type,
                "group_id": group_map.get(entry["group_code"]),
                "reconcile": bool(entry.get("reconcile")),
                "deprecated": bool(entry.get("deprecated")),
            }
        return account_map, account_data

    def _prepare_template_section(
        self,
        template_code,
        base_template_values,
        base_account_codes,
        account_map,
    ):
        template_values = {}
        for key, value in (base_template_values or {}).items():
            if key == "code_digits":
                continue
            mapped_value = self._map_template_value(
                key,
                value,
                base_account_codes,
                account_map,
            )
            if mapped_value is not None:
                template_values[key] = mapped_value
        template_values.setdefault("name", template_code)
        template_values["code_digits"] = 1
        return template_values

    def _map_template_value(
        self,
        key,
        value,
        base_account_codes,
        account_map,
    ):
        if isinstance(value, str):
            mapped = self._map_account_from_base(
                value,
                base_account_codes,
                account_map,
            )
            if mapped:
                return mapped
        if isinstance(value, dict):
            return {
                key_lang: self._map_template_value(
                    key,
                    val,
                    base_account_codes,
                    account_map,
                )
                for key_lang, val in value.items()
            }
        if isinstance(value, list):
            return [
                self._map_template_value(
                    key,
                    item,
                    base_account_codes,
                    account_map,
                )
                for item in value
            ]
        return value

    def _prepare_company_section(
        self,
        company_values,
        base_account_codes,
        account_map,
    ):
        values = copy.deepcopy(company_values or {})
        for field, current in list(values.items()):
            if not current:
                continue
            if field.endswith("_tax_id"):
                values[field] = False
                continue
            if field.endswith("_account_id") or field.startswith("property_account_"):
                values[field] = self._map_company_default(
                    current,
                    base_account_codes,
                    account_map,
                )
        return values

    def _map_company_default(self, value, base_account_codes, account_map):
        if not value:
            return False
        if isinstance(value, str):
            return self._map_account_from_base(value, base_account_codes, account_map) or False
        if isinstance(value, (int, bool)):
            return value
        if isinstance(value, dict):
            return {
                key: self._map_company_default(val, base_account_codes, account_map)
                for key, val in value.items()
            }
        if isinstance(value, (list, tuple)):
            return [
                self._map_company_default(item, base_account_codes, account_map)
                for item in value
            ]
        return value

    def _prepare_reference_models(
        self,
        base_data,
        base_account_codes,
        account_map,
        template_code,
    ):
        tax_group_records = copy.deepcopy((base_data or {}).get("account.tax.group", {}))
        tax_group_data, tax_group_map = self._prepare_tax_group_records(
            tax_group_records,
            base_account_codes,
            account_map,
            template_code,
        )
        taxes = self._prepare_tax_records(
            copy.deepcopy((base_data or {}).get("account.tax", {})),
            base_account_codes,
            account_map,
            tax_group_map,
        )
        journals = self._prepare_journal_records(
            copy.deepcopy((base_data or {}).get("account.journal", {})),
            base_account_codes,
            account_map,
        )
        fiscal_positions = self._prepare_fiscal_position_records(
            copy.deepcopy((base_data or {}).get("account.fiscal.position", {})),
            base_account_codes,
            account_map,
        )
        reconcile_models = self._prepare_reconcile_model_records(
            copy.deepcopy((base_data or {}).get("account.reconcile.model", {})),
            base_account_codes,
            account_map,
        )
        return {
            "account.tax.group": tax_group_data,
            "account.tax": taxes,
            "account.journal": journals,
            "account.fiscal.position": fiscal_positions,
            "account.reconcile.model": reconcile_models,
        }

    def _prepare_tax_group_records(self, tax_group_records, base_account_codes, account_map, template_code):
        cleaned = {}
        mapping = {}
        used_identifiers = set()
        for old_xmlid, values in (tax_group_records or {}).items():
            entry = copy.deepcopy(values)
            for field in ("tax_payable_account_id", "tax_receivable_account_id"):
                if field in entry:
                    mapped = self._map_account_from_base(
                        entry[field],
                        base_account_codes,
                        account_map,
                    )
                    entry[field] = mapped if mapped else False
            identifier_seed = f"{template_code}_{old_xmlid}"
            new_identifier = self._make_identifier("taxgrp", identifier_seed, used_identifiers)
            xmlid = f"l10n_ec_coa_import.{new_identifier}"
            cleaned[xmlid] = entry

            module = False
            bare_name = old_xmlid
            if isinstance(old_xmlid, str) and "." in old_xmlid:
                module, bare_name = old_xmlid.split(".", 1)

            aliases = {
                old_xmlid,
                bare_name,
                f"account.tax.group,{old_xmlid}",
                f"account.tax.group,{bare_name}",
            }
            if module:
                aliases.add(f"{module}.{bare_name}")
                aliases.add(f"account.tax.group,{module}.{bare_name}")
            else:
                aliases.add(f"l10n_ec.{bare_name}")

            for alias in aliases:
                mapping[alias] = xmlid
        return cleaned, mapping

    def _prepare_tax_records(self, tax_records, base_account_codes, account_map, tax_group_map):
        cleaned = {}
        for xmlid, values in (tax_records or {}).items():
            entry = copy.deepcopy(values)
            for field in (
                "cash_basis_account_id",
                "cash_basis_base_account_id",
                "cash_basis_transition_account_id",
            ):
                if field in entry:
                    mapped = self._map_account_from_base(
                        entry[field],
                        base_account_codes,
                        account_map,
                    )
                    entry[field] = mapped if mapped else False
            for field in (
                "repartition_line_ids",
                "invoice_repartition_line_ids",
                "refund_repartition_line_ids",
            ):
                if field in entry:
                    entry[field] = self._map_commands_accounts(
                        entry[field],
                        base_account_codes,
                        account_map,
                        drop_on_missing=False,
                    )
            mapped_group = self._map_tax_group_reference(entry.get("tax_group_id"), tax_group_map)
            if not mapped_group:
                tax_label = entry.get("name")
                if isinstance(tax_label, dict):
                    tax_label = tax_label.get("en_US") or next(iter(tax_label.values()), xmlid)
                raise UserError(
                    _(
                        "Unable to map tax group reference '%s' while preparing tax '%s'.",
                        entry.get("tax_group_id"),
                        tax_label or xmlid,
                    )
                )
            entry["tax_group_id"] = mapped_group
            cleaned[xmlid] = entry
        return cleaned

    def _map_tax_group_reference(self, reference, tax_group_map):
        if not reference:
            return False
        candidates = []
        if isinstance(reference, str):
            candidates.append(reference)
        elif isinstance(reference, (list, tuple)):
            str_parts = [part for part in reference if isinstance(part, str)]
            if len(str_parts) == 2:
                candidates.append(",".join(str_parts))
                candidates.append(str_parts[1])
            candidates.extend(str_parts)
        elif isinstance(reference, dict):
            for key in ("xmlid", "xml_id", "name", "id"):
                value = reference.get(key)
                if isinstance(value, str):
                    candidates.append(value)
        for candidate in candidates:
            if not candidate:
                continue
            mapped = tax_group_map.get(candidate)
            if mapped:
                return mapped
            if candidate.startswith("account.tax.group,"):
                bare = candidate.split(",", 1)[1]
                mapped = tax_group_map.get(bare)
                if mapped:
                    return mapped
            if "." in candidate:
                _, bare = candidate.split(".", 1)
                mapped = tax_group_map.get(bare)
                if mapped:
                    return mapped
        return False

    def _prepare_journal_records(self, journal_records, base_account_codes, account_map):
        cleaned = {}
        account_fields = {
            "default_account_id",
            "refund_account_id",
            "suspense_account_id",
            "profit_account_id",
            "loss_account_id",
            "payment_debit_account_id",
            "payment_credit_account_id",
        }
        for xmlid, values in (journal_records or {}).items():
            entry = copy.deepcopy(values)
            for field in list(account_fields):
                if field in entry:
                    mapped = self._map_account_from_base(
                        entry[field],
                        base_account_codes,
                        account_map,
                    )
                    if mapped:
                        entry[field] = mapped
                    else:
                        entry.pop(field, None)
            cleaned[xmlid] = entry
        return cleaned

    def _prepare_fiscal_position_records(
        self,
        fiscal_records,
        base_account_codes,
        account_map,
    ):
        cleaned = {}
        for xmlid, values in (fiscal_records or {}).items():
            entry = copy.deepcopy(values)
            if "account_ids" in entry:
                entry["account_ids"] = self._map_commands_accounts(
                    entry["account_ids"],
                    base_account_codes,
                    account_map,
                    drop_on_missing=True,
                )
            cleaned[xmlid] = entry
        return cleaned

    def _prepare_reconcile_model_records(
        self,
        reconcile_records,
        base_account_codes,
        account_map,
    ):
        cleaned = {}
        for xmlid, values in (reconcile_records or {}).items():
            entry = copy.deepcopy(values)
            if "account_id" in entry:
                mapped = self._map_account_from_base(
                    entry["account_id"],
                    base_account_codes,
                    account_map,
                )
                if mapped:
                    entry["account_id"] = mapped
                else:
                    entry.pop("account_id", None)
            if "line_ids" in entry:
                entry["line_ids"] = self._map_commands_accounts(
                    entry["line_ids"],
                    base_account_codes,
                    account_map,
                    drop_on_missing=True,
                )
            cleaned[xmlid] = entry
        return cleaned

    def _map_commands_accounts(
        self,
        commands,
        base_account_codes,
        account_map,
        drop_on_missing,
    ):
        mapped = []
        for command in commands or []:
            if isinstance(command, tuple) and len(command) >= 3 and isinstance(command[2], dict):
                command_parts = list(command)
                values = dict(command_parts[2])
                remove_line = False
                for field, current in list(values.items()):
                    if field.endswith("account_id") and isinstance(current, str):
                        mapped_value = self._map_account_from_base(
                            current,
                            base_account_codes,
                            account_map,
                        )
                        if mapped_value:
                            values[field] = mapped_value
                        elif drop_on_missing:
                            remove_line = True
                            break
                        else:
                            values[field] = False
                if remove_line:
                    continue
                command_parts[2] = values
                mapped.append(tuple(command_parts))
            else:
                mapped.append(command)
        return mapped

    def _map_account_from_base(self, reference, base_account_codes, account_map):
        if not isinstance(reference, str):
            return reference
        code = base_account_codes.get(reference)
        if not code:
            return False
        return account_map.get(code)

    def _restore_account_codes(self, company, account_payload):
        if not account_payload:
            return
        Account = self.env["account.account"].sudo().with_context(active_test=False)
        for xmlid, values in (account_payload or {}).items():
            target_code = values.get("code")
            if not target_code:
                continue
            record = self.env.ref(xmlid, raise_if_not_found=False)
            if not record:
                record = self.env.ref(f"l10n_ec_coa_import.{xmlid}", raise_if_not_found=False)
            if not record:
                continue
            account = Account.browse(record.id)
            if not account or company not in account.company_ids:
                continue
            if account.code != target_code:
                _logger.info(
                    "Restoring custom COA account %s (%s) code from %s to %s",
                    account.display_name,
                    account.id,
                    account.code,
                    target_code,
                )
                account.write({"code": target_code})

    def _make_identifier(self, prefix, value, existing):
        base = re.sub(r"[^0-9a-zA-Z]+", "_", (value or "").lower()).strip("_")
        if not base:
            base = "line"
        base = base[:40]
        candidate = f"{prefix}_{base}"
        for idx in count(1):
            if candidate not in existing:
                break
            candidate = f"{prefix}_{base}_{idx}"
        existing.add(candidate)
        return candidate

    def _determine_code_digits(self, accounts):
        numeric_lengths = []
        for entry in accounts:
            digits = "".join(ch for ch in entry["code"] if ch.isdigit())
            if digits:
                numeric_lengths.append(len(digits))
        return max(numeric_lengths or [6])

    def _generate_template_name(self, company):
        timestamp = fields.Datetime.now().strftime("%Y-%m-%d %H:%M")
        return _("%s Custom COA (%s)") % (company.display_name, timestamp)

    def _clone_localization_records(self, company):
        """Placeholder for cloning taxes, journals and other localization data.

        TODO: Implement record duplication from l10n_ec and l10n_ec_base while keeping
        references to the original templates.
        """
        _logger.info(
            "Localization cloning not yet implemented for company %s.",
            company.display_name,
        )

    # ------------------------------------------------------------------
    # Onboarding utilities (placeholder for template download, to be implemented)
    # ------------------------------------------------------------------
